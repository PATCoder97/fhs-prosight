"""
Service layer for evaluation data management.

This module provides business logic for:
1. Excel file upload and import with upsert logic
2. Evaluation search with filters and pagination
3. Response transformation (flat DB rows → nested evaluation groups)
"""

import logging
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.exc import IntegrityError
from fastapi import HTTPException
import openpyxl

from app.models.evaluation import Evaluation

logger = logging.getLogger(__name__)


# Excel column mapping (Chinese headers → English database fields)
EXCEL_COLUMN_MAPPING = {
    "評核年月": "term_code",
    "工號": "employee_id",
    "姓名": "employee_name",
    "職等": "job_level",
    "國籍": "nation",
    "部門代碼": "dept_code",
    "部門名稱": "dept_name",
    "職等六碼": "grade_code",
    "職等名稱": "grade_name",
    "初核成績": "init_score",
    "初核評語": "init_comment",
    "初核主管": "init_reviewer",
    "複核成績": "review_score",
    "複核評語": "review_comment",
    "複核主管": "review_reviewer",
    "核定成績": "final_score",
    "核定評語": "final_comment",
    "核定主管": "final_reviewer",
    "經理室初核成績": "mgr_init_score",
    "經理室初核評語": "mgr_init_comment",
    "經理室初核主管": "mgr_init_reviewer",
    "經理室複核成績": "mgr_review_score",
    "經理室複核評語": "mgr_review_comment",
    "經理室複核主管": "mgr_review_reviewer",
    "經理室核定成績": "mgr_final_score",
    "經理室核定評語": "mgr_final_comment",
    "經理室核定主管": "mgr_final_reviewer",
    "請假總日數": "leave_days",
}


async def upload_evaluations_from_excel(
    db: AsyncSession,
    file_path: str
) -> dict:
    """
    Parse Excel file and import/update evaluation records.

    Args:
        db: Database session
        file_path: Path to Excel file (.xlsx)

    Returns:
        dict with upload summary:
        {
            "success": True,
            "summary": {
                "total_rows": 150,
                "created": 120,
                "updated": 30,
                "errors": 0
            },
            "error_details": []
        }

    Raises:
        HTTPException(400): Invalid file format or missing required columns
        HTTPException(500): Unexpected processing error
    """
    logger.info(f"Starting Excel upload from: {file_path}")

    # Initialize counters
    total_rows = 0
    created_count = 0
    updated_count = 0
    error_count = 0
    error_details = []

    try:
        # Open Excel file (read-only mode for performance)
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active

        # Read header row
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell else "" for cell in header_row]

        # Validate required columns exist
        required_columns = ["評核年月", "工號"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )

        # Create column index mapping
        column_indices = {}
        for chinese_name, english_name in EXCEL_COLUMN_MAPPING.items():
            if chinese_name in headers:
                column_indices[english_name] = headers.index(chinese_name)

        logger.info(f"Mapped {len(column_indices)} columns from Excel")

        # Collect all rows first, then separate into inserts and updates
        all_rows_data = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1

            try:
                # Extract data from row using column indices
                row_data = {}
                for english_name, col_idx in column_indices.items():
                    value = row[col_idx] if col_idx < len(row) else None

                    # Convert value to string (except leave_days which is float)
                    if value is not None:
                        if english_name == "leave_days":
                            try:
                                row_data[english_name] = float(value)
                            except (ValueError, TypeError):
                                row_data[english_name] = None
                        else:
                            row_data[english_name] = str(value).strip()
                    else:
                        row_data[english_name] = None

                # Validate required fields
                if not row_data.get("term_code") or not row_data.get("employee_id"):
                    error_count += 1
                    error_details.append({
                        "row": row_num,
                        "error": "Missing required fields: term_code or employee_id"
                    })
                    continue

                all_rows_data.append((row_num, row_data))

            except Exception as e:
                error_count += 1
                error_details.append({
                    "row": row_num,
                    "error": str(e)
                })
                logger.warning(f"Row {row_num} parsing error: {e}")
                continue

        logger.info(f"Parsed {len(all_rows_data)} valid rows from Excel")

        # Fetch all existing records in one query
        existing_keys = [(r[1]["term_code"], r[1]["employee_id"]) for r in all_rows_data]

        if existing_keys:
            # Build query to fetch all existing records
            conditions = []
            for term_code, employee_id in existing_keys:
                conditions.append(
                    (Evaluation.term_code == term_code) &
                    (Evaluation.employee_id == employee_id)
                )

            # Use OR to combine all conditions
            from sqlalchemy import or_
            stmt = select(Evaluation).where(or_(*conditions))
            result = await db.execute(stmt)
            existing_records = result.scalars().all()

            # Create a map of existing records
            existing_map = {
                (rec.term_code, rec.employee_id): rec
                for rec in existing_records
            }

            logger.info(f"Found {len(existing_map)} existing records to update")
        else:
            existing_map = {}

        # Separate into inserts and updates
        records_to_insert = []
        records_to_update = []

        for row_num, row_data in all_rows_data:
            key = (row_data["term_code"], row_data["employee_id"])

            if key in existing_map:
                # UPDATE existing record
                existing_record = existing_map[key]
                for field, value in row_data.items():
                    setattr(existing_record, field, value)
                records_to_update.append(existing_record)
                updated_count += 1
            else:
                # INSERT new record
                new_record = Evaluation(**row_data)
                records_to_insert.append(new_record)
                created_count += 1

        # Bulk insert new records
        if records_to_insert:
            db.add_all(records_to_insert)
            logger.info(f"Bulk inserting {len(records_to_insert)} new records")

        # Commit all changes (updates are already tracked in session)
        try:
            await db.commit()
            logger.info(f"Upload complete: {created_count} created, {updated_count} updated, {error_count} errors")
        except Exception as commit_error:
            await db.rollback()
            logger.error(f"Commit failed: {commit_error}", exc_info=True)
            raise

        return {
            "success": True,
            "summary": {
                "total_rows": total_rows,
                "created": created_count,
                "updated": updated_count,
                "errors": error_count
            },
            "error_details": error_details
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel upload failed: {e}", exc_info=True)
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process Excel file: {str(e)}"
        )


async def search_evaluations(
    db: AsyncSession,
    employee_id: Optional[str] = None,
    term_code: Optional[str] = None,
    dept_code: Optional[str] = None,
    page: int = 1,
    page_size: int = 50
) -> dict:
    """
    Search evaluation records with filters and pagination.

    Args:
        db: Database session
        employee_id: Exact match filter (e.g., 'VNW0018983')
        term_code: Exact match filter (e.g., '25B')
        dept_code: Prefix match filter (e.g., '78' matches '7800', '7810')
        page: Page number (1-indexed, default: 1)
        page_size: Items per page (default: 50, max: 100)

    Returns:
        dict matching SearchResponse schema:
        {
            "total": 250,
            "page": 1,
            "page_size": 50,
            "results": [...]
        }

    Raises:
        HTTPException(422): Invalid query parameters
    """
    logger.info(f"Searching evaluations: employee_id={employee_id}, term_code={term_code}, dept_code={dept_code}, page={page}, page_size={page_size}")

    # Validate pagination parameters
    if page < 1:
        raise HTTPException(
            status_code=422,
            detail="Page number must be >= 1"
        )

    if page_size < 1 or page_size > 100:
        raise HTTPException(
            status_code=422,
            detail="Page size must be between 1 and 100"
        )

    try:
        # Build base query
        query = select(Evaluation)

        # Apply filters
        if employee_id:
            query = query.where(Evaluation.employee_id == employee_id)

        if term_code:
            query = query.where(Evaluation.term_code == term_code)

        if dept_code:
            # Prefix match (LIKE 'dept_code%')
            query = query.where(Evaluation.dept_code.like(f"{dept_code}%"))

        # Count total records (before pagination)
        count_query = select(func.count()).select_from(query.subquery())
        count_result = await db.execute(count_query)
        total = count_result.scalar()

        # Apply pagination
        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        # Execute query
        result = await db.execute(query)
        rows = result.scalars().all()

        # Transform results to nested structure
        results = []
        for row in rows:
            evaluation_dict = {
                "id": row.id,
                "term_code": row.term_code,
                "employee_id": row.employee_id,
                "employee_name": row.employee_name,
                "job_level": row.job_level,
                "nation": row.nation,
                "dept_code": row.dept_code,
                "dept_name": row.dept_name,
                "grade_code": row.grade_code,
                "grade_name": row.grade_name,
                "dept_evaluation": {
                    "init": {
                        "score": row.init_score,
                        "comment": row.init_comment,
                        "reviewer": row.init_reviewer
                    },
                    "review": {
                        "score": row.review_score,
                        "comment": row.review_comment,
                        "reviewer": row.review_reviewer
                    },
                    "final": {
                        "score": row.final_score,
                        "comment": row.final_comment,
                        "reviewer": row.final_reviewer
                    }
                },
                "mgr_evaluation": {
                    "init": {
                        "score": row.mgr_init_score,
                        "comment": row.mgr_init_comment,
                        "reviewer": row.mgr_init_reviewer
                    },
                    "review": {
                        "score": row.mgr_review_score,
                        "comment": row.mgr_review_comment,
                        "reviewer": row.mgr_review_reviewer
                    },
                    "final": {
                        "score": row.mgr_final_score,
                        "comment": row.mgr_final_comment,
                        "reviewer": row.mgr_final_reviewer
                    }
                },
                "leave_days": row.leave_days,
                "created_at": row.created_at,
                "updated_at": row.updated_at
            }
            results.append(evaluation_dict)

        logger.info(f"Found {total} total records, returning page {page} with {len(results)} results")

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Search failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to search evaluations: {str(e)}"
        )
